import openpyxl
import sys

def analyze_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    print(f"Workbook: {file_path}")
    print(f"Sheet names: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        
        # Print dimensions
        print(f"Dimensions: {ws.dimensions}")
        
        # Iterate over rows to find content
        print("Content Preview (first 100 rows):")
        for i, row in enumerate(ws.iter_rows(values_only=False)):
            if i > 100: break
            row_data = []
            has_data = False
            for cell in row:
                val = cell.value
                if val is not None:
                    has_data = True
                    # Check for formula
                    if isinstance(val, str) and val.startswith('='):
                        row_data.append(f"FORMULA: {val}")
                    else:
                        row_data.append(str(val))
                else:
                    row_data.append("")
            
            if has_data:
                # Filter out empty strings from the end of the list to make it cleaner
                while row_data and row_data[-1] == "":
                    row_data.pop()
                print(f"Row {i+1}: {row_data}")
                
        # Check for comments
        print("\nComments:")
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    print(f"Cell {cell.coordinate}: {cell.comment.text}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        analyze_excel(sys.argv[1])
    else:
        print("Please provide file path")
